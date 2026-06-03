import openpyxl
wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)
ws = wb['Tables']

# Army Tactics section starts around row 16-32
# Columns: Name, Min Level, Type(s), Tac Actions Granted, Trait, Description
print("=== Army Tactics (rows 16-32) ===")
for i in range(16, 33):
    row = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    if any(v is not None for v in row):
        # Extract columns after the army stats section (starting around col 38 / 0-indexed 37)
        # The tactics table starts at around col 37 (AK in spreadsheet terms)
        print(f"Row {i}: {row}")

print()
print("=== Specialized Army Modifiers (rows 26-32+) ===")
for i in range(26, 34):
    row = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    if any(v is not None for v in row):
        print(f"Row {i}: {row}")

print()
print("=== Army unit rows (basic armies + modifiers) ===")
# Row 4 onwards in the "Basic Army" section
for i in range(4, 10):
    row = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    if any(v is not None for v in row):
        print(f"Row {i}: {row}")
