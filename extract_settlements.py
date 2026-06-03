import openpyxl
wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)

# List all sheet names
print('Sheet names:', wb.sheetnames)
print()

ws = wb['Settlements']
print(f'=== SETTLEMENTS SHEET ===')
print(f'Dimensions: {ws.dimensions}')
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
    for cell in row:
        if cell.value is not None:
            print(f'  {cell.coordinate}: {repr(cell.value)}')
