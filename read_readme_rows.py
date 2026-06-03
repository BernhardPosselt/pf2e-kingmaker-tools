import openpyxl
wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)
ws = wb['README']

# Just output the raw README as structured text with row numbers
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
    vals = []
    for cell in row:
        v = cell.value
        if v is not None:
            vals.append(f"[{cell.coordinate}]{v}")
    if vals:
        # Print with row number
        row_num = row[0].row
        print(f"ROW {row_num:3d}: {' | '.join(vals)}")
