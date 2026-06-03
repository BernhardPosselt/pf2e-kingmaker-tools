import openpyxl
import json

wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)
ws = wb['Urban Grid Template']

print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print()

# Print all cells with values
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
    for cell in row:
        if cell.value is not None:
            print(f"  {cell.coordinate}: {repr(cell.value)}")
