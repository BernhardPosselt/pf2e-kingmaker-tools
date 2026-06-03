import openpyxl
wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)

# Read all sheets that contain structured data
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name} | Dimensions: {ws.dimensions} | Rows: {ws.max_row} | Cols: {ws.max_column}")
    print('='*80)
    
    # Print first 50 rows to understand structure
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), max_col=ws.max_column, values_only=False):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                vals.append(f"[{cell.coordinate}]{v}")
        if vals:
            print(" | ".join(vals))
            count += 1
    if ws.max_row > 60:
        print(f"... (truncated, {ws.max_row} total rows)")
