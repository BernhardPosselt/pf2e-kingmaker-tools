import openpyxl
wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)
ws = wb['Tables']

# Extract the Charters, Governments, Skills, and Leadership sections properly
print("=== R40-R63: Charters/Gov section ===")
for row in ws.iter_rows(min_row=40, max_row=65, max_col=ws.max_column, values_only=False):
    vals = []
    for cell in row:
        v = cell.value
        if v is not None:
            vals.append(f"[{cell.coordinate}]{v}")
    if vals:
        row_num = row[0].row
        print(f"R{row_num}: {' | '.join(vals[:5])}")  # first 5 cols

print("\n=== R64-R74: Leadership section ===")
for row in ws.iter_rows(min_row=64, max_row=74, max_col=ws.max_column, values_only=False):
    vals = []
    for cell in row:
        v = cell.value
        if v is not None:
            vals.append(f"[{cell.coordinate}]{v}")
    if vals:
        row_num = row[0].row
        print(f"R{row_num}: {' | '.join(vals[:5])}")

print("\n=== R35-R39: Ability Scores ===")
for row in ws.iter_rows(min_row=35, max_row=39, max_col=ws.max_column, values_only=False):
    vals = []
    for cell in row:
        v = cell.value
        if v is not None:
            vals.append(f"[{cell.coordinate}]{v}")
    if vals:
        row_num = row[0].row
        print(f"R{row_num}: {' | '.join(vals[:3])}")
