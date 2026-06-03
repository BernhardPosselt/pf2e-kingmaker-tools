import openpyxl
wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)
ws = wb['Tables']

# Read rows 4-12 for basic army units (Army Template Names, Type, etc.)
# These are the rows from the "Basic Armies" section
print("=== Army units with full stats (rows 4-12) ===")
for i in range(4, 13):
    row = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    if any(v is not None for v in row[:43]):
        # cols 28-38ish: Army name, type, consumption, HP, min level, attacks, ranged Ammo, maneuver save, special faction, accessible, starting tactics, description
        print(f"Row {i}: {row}")

print()
print("=== Kingdom Feats section (rows 75-120ish) ===")
found_feat_headers = False
for i in range(75, 95):
    row = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    if any(v is not None for v in row[:7]):
        print(f"Row {i}: {row}")
        found_feat_headers = True

print()
print("=== Specialized Army Modifiers table (columns 21-28, rows 27-34) ===")
# The "Specialized Army Modifiers" is at row 25, with headers at row 26
for i in range(25, 35):
    row = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    # Focus on cols 19-27 (the modifier section)
    modifier_cols = row[19:28] if len(row) > 28 else row
    print(f"Row {i} [modifiers]: {modifier_cols}")
