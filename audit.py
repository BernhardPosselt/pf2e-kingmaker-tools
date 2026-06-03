#!/usr/bin/env python3
"""
Audit workbook README sheet against repo code and JSON.
Extracts all data references from the README and checks them against:
1. Kotlin source files
2. JSON data files
"""
import os, json, re, glob

REPO = "/home/grego/code/pf2e-kingmaker-tools"

# ---- Read README sheet ----
import openpyxl
wb = openpyxl.load_workbook(f'{REPO}/Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)
ws = wb['README']
readme_text = ""
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
    for cell in row:
        if cell:
            readme_text += str(cell) + "\n"

# ---- Catalog repo files ----
json_files = []
kt_files = []
for root, dirs, files in os.walk(REPO):
    if '.git' in root: continue
    for f in files:
        fp = os.path.join(root, f)
        if f.endswith('.json'):
            json_files.append(fp)
        elif f.endswith('.kt') or f.endswith('.kts'):
            kt_files.append(fp)

# Read all JSON data (structure names, event names, etc.)
json_names = set()
json_data = {}
for jf in json_files:
    bn = os.path.basename(jf).replace('.json', '')
    json_names.add(bn)
    try:
        with open(jf) as f:
            data = json.load(f)
        json_data[bn] = data
    except:
        pass

# Read all Kotlin source
kt_text = ""
for kf in kt_files:
    with open(kf, errors='replace') as f:
        kt_text += f.read() + "\n"

# ---- Extract structured items from the README ----
# Items to audit
audit_items = []

# 1. Sheets mentioned in README
sheets_in_changelog = []
for line in readme_text.split('\n'):
    # "KingdomSheets" range
    if 'KingdomSheets' in line:
        audit_items.append(('Changelog', 'KingdomSheets range', line.strip()))
    # "Structures" tab
    if "Structures" in line:
        audit_items.append(('Changelog', 'Structures tab reference', line.strip()))
    # "Urban Grids" / "Settlements" tabs
    if "Urban Grids" in line:
        audit_items.append(('Changelog', 'Urban Grids tab reference', line.strip()))
    if "Settlements" in line:
        audit_items.append(('Changelog', 'Settlements tab reference', line.strip()))
    if "'⚙️'" in line or "⚙️" in line:
        audit_items.append(('Changelog', '⚙️ settings tab', line.strip()))
    if "History" in line:
        audit_items.append(('Changelog', 'History tab reference', line.strip()))
    if "Turn Tracker" in line:
        audit_items.append(('Changelog', 'Turn Tracker tab reference', line.strip()))

# 2. Features/items from changelog
features = [
    # CURRENT
    ("Inspiring Entertainment feat", "Culture skills"),
    ("Practical Magic feat", "Culture skills / Magic checks"),
    ("Sewer System consumption reduction", "consumption"),
    ("Mill water-adjacent -1 Consumption", "water-adjacent Mills"),
    ("ACTIVITY_MODIFIERS_FOR_SETTLEMENT function", "structure item bonus"),
    ("CircumstancePenalty boxes", "Agriculture, Arts, Defense, Exploration, Folklore, Magic, Scholarship, Wilderness"),
    ("Base Image URL column", "Structures tab"),
    ("Edifice structure trait", "Urban Grids"),
    ("Stockyard consumption reduction", "consumption"),
    ("Army tracking tab", "armies"),
    ("Construction skill modifier", "Build Structure action"),
    ("Farmlands consumption (not Food Commodities)", "Gain resource step"),
    ("Clandestine business DC calculation", "clandestine business"),
    ("TurnModifier / RawModifier", "Turn Tracker modifiers"),
    ("V&K Practical Magic nerf", "configurable settings"),
    ("Feat status bonuses", "fix"),
    ("Warden penalty", "penalty"),
    ("⚙️ Tab homebrew", "settings"),
    ("Untrained Skill Bonus setting", "settings"),
    ("RP To XP Conversion Rate setting", "settings"),
    ("High level structure item bonus bug", "settlement level"),
    ("Available items levels in settlement", "Settlements tab"),
    ("XP-related settings", "Turn Tracker"),
    ("Resource Collection Work Sites", "Turn Tracker"),
    ("Commodity Storage by Kingdom Size", "calculation"),
    ("Block size range (Town=1 block)", "settlement type"),
    ("Minimal Capital Influence setting", "Farmlands"),
    ("Skill Training comma-separated", "Boating, Defense"),
    ("Settlement type by block size", "determination"),
    ("XP auto-decrease on level up", "History tab"),
    ("XP To Level setting", "VanceMadrox"),
    ("Mill/Stockyard/Sewer System consumption update", "Consumption"),
    ("Expansion Expert bonus Claim Hex missing", "known issue"),
    # 1.1.0
    ("Trained/expert/master skill gates", "Turn Tracker N/A"),
    ("Army Activities initial", "armies"),
    ("Clear Values button", "initial values"),
    ("Unrest penalties to skill checks", "bug fix"),
    ("Fame/Infamy auto-add", "new Turn"),
    ("Trade modifier Collect Taxes", "Trade"),
    ("Building on Rough Terrain table", "Establish Worksite"),
    # 1.0.0
    ("First feature complete", "v1.0.0"),
]

aligned = []
gaps = []
orphans = []

print("=" * 100)
print("AUDIT: README sheet vs repo code/JSON")
print("=" * 100)

# Check each feature against code/JSON
for feature, context in features:
    found_in_code = feature.split()[0].lower() in kt_text.lower()
    found_in_json = False
    for jn in json_names:
        if feature.lower() in jn.lower():
            found_in_json = True
            break
    # Check README text exists
    in_readme = feature in readme_text
    
    if found_in_code or found_in_json:
        aligned.append(f"  [ALIGNED] {feature} ({context})")
    else:
        # Check partial match in code
        partial = False
        for word in feature.split()[:3]:
            if len(word) > 4 and word.lower() in kt_text.lower():
                partial = True
                break
        if partial:
            aligned.append(f"  [ALIGNED*] {feature} ({context}) [partial code match]")
        else:
            gaps.append(f"  [GAP] {feature} ({context}) - mentioned in README changelog but NOT found clearly in code/JSON")

# Now check: things in code/JSON that are NOT in README
# Kingdom skills in code
with open(f'{REPO}/src/commonMain/kotlin/at/posselt/pfrpg2e/data/kingdom/KingdomSkill.kt') as f:
    skill_content = f.read()

# Check for structure traits
with open(f'{REPO}/src/commonMain/kotlin/at/posselt/pfrpg2e/data/kingdom/structures/StructureTrait.kt') as f:
    trait_content = f.read()

# Extract enum values from code
kt_enums = re.findall(r'(?:enum class|object)\s+(\w+)', kt_text)
kt_classes = re.findall(r'(?:class|object|data class|sealed class)\s+(\w+)', kt_text)

# Structures in JSON but not in README
structure_names = set(sorted(json_names))
# Filter to just structure files
structure_files = sorted([os.path.basename(j).replace('.json','') for j in json_files if '/structures/' in j])

# Events
event_files = sorted([os.path.basename(j).replace('.json','') for j in json_files if '/events/' in j])
charter_files = sorted([os.path.basename(j).replace('.json','') for j in json_files if '/charters/' in j])
government_files = sorted([os.path.basename(j).replace('.json','') for j in json_files if '/governments/' in j])
milestone_files = sorted([os.path.basename(j).replace('.json','') for j in json_files if '/milestones/' in j])
camping_activity_files = sorted([os.path.basename(j).replace('.json','') for j in json_files if '/camping-activities/' in j])

print("\n--- STRUCTURE DATA FILES ---")
for s in structure_files:
    in_readme = s in readme_text
    status = "ALIGNED" if in_readme else "ORPHAN (in JSON, not in README)"
    print(f"  [{status}] {s}")

print(f"\n--- EVENT DATA FILES ({len(event_files)} total) ---")
for e in event_files:
    in_readme = e in readme_text
    status = "ALIGNED" if in_readme else "ORPHAN (in JSON, not in README)"
    print(f"  [{status}] {e}")

print(f"\n--- CHARTER DATA FILES ---")
for c in charter_files:
    in_readme = c in readme_text
    status = "ALIGNED" if in_readme else "ORPHAN (in JSON, not in README)"
    print(f"  [{status}] {c}")

print(f"\n--- GOVERNMENT DATA FILES ---")
for g in government_files:
    in_readme = g in readme_text
    status = "ALIGNED" if in_readme else "ORPHAN (in JSON, not in README)"
    print(f"  [{status}] {g}")

print(f"\n--- CAMPING ACTIVITY DATA FILES ---")
for ca in camping_activity_files:
    in_readme = ca in readme_text
    status = "ALIGNED" if in_readme else "ORPHAN (in JSON, not in README)"
    print(f"  [{status}] {ca}")

print(f"\n--- MILESTONE DATA FILES ---")
for m in milestone_files:
    in_readme = m in readme_text
    status = "ALIGNED" if in_readme else "ORPHAN (in JSON, not in README)"
    print(f"  [{status}] {m}")

# Check Kingdom Skills
print("\n--- KINGDOM SKILLS ---")
skill_match = re.findall(r'(\w+)\s*\(', skill_content)
skills_in_code = []
for line in skill_content.split('\n'):
    line = line.strip()
    if line and not line.startswith('//') and not line.startswith('*') and not line.startswith('package') and not line.startswith('import'):
        if '(' in line and ')' in line:
            name = line.split('(')[0].strip()
            if name[0].isupper() and 'object' not in name and 'enum' not in name:
                skills_in_code.append(name)
                
print("Skills in code:", skills_in_code)

# README mentions these skills
readme_skills = ["Agriculture", "Arts", "Defense", "Exploration", "Folklore", "Magic", "Scholarship", "Wilderness"]
for sk in readme_skills:
    found = sk in skill_content
    if found:
        print(f"  [ALIGNED] {sk} skill")
    else:
        print(f"  [GAP] {sk} skill - mentioned in README but not found in code")

print("\n\n" + "=" * 100)
print("SUMMARY: Changelog Features - Aligned vs Gaps")
print("=" * 100)
for a in aligned:
    print(a)
for g in gaps:
    print(g)

print(f"\nTotals: {len(aligned)} aligned, {len(gaps)} gaps")
print(f"JSON data files: {len(structure_files)} structures, {len(event_files)} events, {len(charter_files)} charters, {len(government_files)} governments, {len(camping_activity_files)} camping activities, {len(milestone_files)} milestones")
