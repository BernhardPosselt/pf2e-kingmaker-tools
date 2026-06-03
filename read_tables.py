import openpyxl
wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)

# Read the Tables sheet - this has the structured data
ws = wb['Tables']
print(f"Tables sheet: {ws.dimensions}, {ws.max_row} rows, {ws.max_column} cols")
print("="*100)

# Find all table headers/sections
current_table = ""
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
    vals = []
    for cell in row:
        v = cell.value
        if v is not None:
            vals.append(str(v))
    if vals:
        line = " | ".join(vals)
        # Check if this looks like a table header
        row_num = row[0].row
        print(f"R{row_num:3d}: {line[:150]}")
