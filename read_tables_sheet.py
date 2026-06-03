import openpyxl
wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)

# Deep dive into the Tables sheet
ws = wb['Tables']
print(f'Tables sheet: {ws.max_row}x{ws.max_column}')
print()

# Print first 30 rows to understand table layout
count = 0
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True), 1):
    # Skip fully empty rows
    if any(v is not None for v in row):
        print(f'Row {i}: {row}')
        count += 1

print()
print('--- Rows 31-80 ---')
for i, row in enumerate(ws.iter_rows(min_row=31, max_row=min(80, ws.max_row), values_only=True), 31):
    if any(v is not None for v in row):
        print(f'Row {i}: {row}')
