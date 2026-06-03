import openpyxl
wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'\n=== {sname} ({ws.max_row}x{ws.max_column}) ===')
    for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
        print(row)
