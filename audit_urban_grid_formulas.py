import openpyxl

wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=False)
ws = wb['Urban Grid Template']

print("=== URBAN GRID TEMPLATE (with formulas) ===")
print()

# Get all cells with values or formulas
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
    for cell in row:
        if cell.value is not None:
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                print(f"  {cell.coordinate}: FORMULA: {val[:200]}")
            else:
                print(f"  {cell.coordinate}: {repr(val)}")
