import openpyxl

wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)
ws = wb['Kingdom Sheet']

print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print()

for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 120), values_only=False):
    for cell in row:
        if cell.value is not None:
            print(f"  {cell.coordinate}: {cell.value!r}")
