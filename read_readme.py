import openpyxl
wb = openpyxl.load_workbook('Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)
ws = wb['README']

print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print("=" * 120)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
    vals = []
    for cell in row:
        v = cell.value
        if v is not None:
            vals.append(f"[{cell.coordinate}]{v}")
    if vals:
        print(" | ".join(vals))
