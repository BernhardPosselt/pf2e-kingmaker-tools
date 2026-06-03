import openpyxl, json, os

REPO = "/home/grego/code/pf2e-kingmaker-tools"
wb = openpyxl.load_workbook(f'{REPO}/Copy of Royal Kingdom Sheet for Pathfinder 2E - CURRENT.xlsx', data_only=True)

# Read Structures sheet to get structure names from workbook
ws_struct = wb['Structures']
structures_in_workbook = set()
for row in ws_struct.iter_rows(min_row=2, max_row=ws_struct.max_row, max_col=1, values_only=True):
    if row[0]:
        structures_in_workbook.add(str(row[0]).strip())

# Read JSON structure files
json_structures = set()
for f in os.listdir(f'{REPO}/data/structures'):
    if f.endswith('.json') and f != 'package.json':
        json_structures.add(f.replace('.json', ''))

# Read Charters
ws_tables = wb['Tables']
charters_in_workbook = set()
for row in ws_tables.iter_rows(min_row=42, max_row=46, max_col=1, values_only=True):
    if row[0]:
        charters_in_workbook.add(str(row[0]).strip())

json_charters = set()
for f in os.listdir(f'{REPO}/data/charters'):
    if f.endswith('.json'):
        json_charters.add(f.replace('.json', ''))

# Read Governments
gov_in_workbook = set()
for row in ws_tables.iter_rows(min_row=57, max_row=62, max_col=1, values_only=True):
    if row[0]:
        gov_in_workbook.add(str(row[0]).strip())

json_gov = set()
for f in os.listdir(f'{REPO}/data/governments'):
    if f.endswith('.json'):
        json_gov.add(f.replace('.json', ''))

# Read Kingdom Skills from Tables sheet
skills_in_workbook = set()
for row in ws_tables.iter_rows(min_row=35, max_row=38, max_col=1, values_only=True):
    if row[0]:
        skills_in_workbook.add(str(row[0]).strip())

# Read Feats from Tables sheet (rows 76-92)
feats_in_workbook = set()
for row in ws_tables.iter_rows(min_row=77, max_row=110, max_col=1, values_only=True):
    if row[0]:
        feats_in_workbook.add(str(row[0]).strip())

# Leadership roles
leaders_in_workbook = set()
for row in ws_tables.iter_rows(min_row=66, max_row=73, max_col=1, values_only=True):
    if row[0]:
        leaders_in_workbook.add(str(row[0]).strip())

print("="*100)
print("STRUCTURE NAMES: Workbook vs JSON")
print("="*100)
print(f"\nStructures in workbook ({len(structures_in_workbook)}):")
for s in sorted(structures_in_workbook):
    in_json = s in json_structures
    print(f"  [{'ALIGNED' if in_json else 'GAP'}] {s}")

print(f"\nStructures in JSON but not in workbook ({len(json_structures - structures_in_workbook)}):")
for s in sorted(json_structures - structures_in_workbook):
    print(f"  [ORPHAN] {s}")

print("\n" + "="*100)
print("CHARTERS: Workbook vs JSON")
print("="*100)
for c in sorted(charters_in_workbook):
    in_json = c in json_charters
    print(f"  [{'ALIGNED' if in_json else 'GAP'}] {c}")
for c in sorted(json_charters - charters_in_workbook):
    print(f"  [ORPHAN in JSON] {c}")

print("\n" + "="*100)
print("GOVERNMENTS: Workbook vs JSON")
print("="*100)
for g in sorted(gov_in_workbook):
    in_json = g in json_gov
    print(f"  [{'ALIGNED' if in_json else 'GAP'}] {g}")
for g in sorted(json_gov - gov_in_workbook):
    print(f"  [ORPHAN in JSON] {g}")

print("\n" + "="*100)
print("KINGDOM SKILLS: Workbook Tables vs Code")
print("="*100)
for s in sorted(skills_in_workbook):
    print(f"  {s}")

print("\n" + "="*100)
print("FEATS from workbook Tables sheet:")
print("="*100)
for f in sorted(feats_in_workbook):
    print(f"  {f}")

print("\n" + "="*100)
print("LEADERSHIP ROLES from workbook Tables sheet:")
print("="*100)
for l in sorted(leaders_in_workbook):
    print(f"  {l}")
